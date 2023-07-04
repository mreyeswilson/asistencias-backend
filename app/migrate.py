
import openpyxl
from models.localidad import Localidad
from models.unidad import Unidad
from db import db

def migrate():
    book = openpyxl.load_workbook("files/TARDIAS.xlsx")

    # remove TODAS from sheetnames variable
    sheetnames = book.sheetnames

    for hoja in sheetnames:
        sheet = book[hoja]
        tecnico = sheet["D2"].value
        localidad = Localidad(hoja, tecnico, "activa")
        db.session.add(localidad)
        db.session.commit()

        for row in range(2, sheet.max_row + 1):
            unidad = sheet.cell(row, 1).value
            unidad = unidad.strip() if unidad else unidad

            if not unidad:
                print(f"No hay unidad en la fila {row} en la hoja {hoja}")
                break

            tipo = sheet.cell(row, 3).value
            codigo = sheet.cell(row, 2).value

            u = Unidad(unidad, codigo, localidad.id, tipo)
            db.session.add(u)

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(e)

    book.close()
    return "Migración finalizada"