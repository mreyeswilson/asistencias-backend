import os
import openpyxl

class Unidad:

    nombre = ""
    codigo = ""
    localidad = ""
    carpeta = ""

    def __init__(self, nombre, codigo, localidad):
        self.nombre = nombre
        self.codigo = codigo
        self.localidad = localidad
        self.carpeta = f"./output/{self.localidad} - {self.nombre}"

    def crear_carpeta(self):
        if not os.path.exists(self.carpeta):
            os.makedirs(self.carpeta)

    def llenar_formato(self):
        book = openpyxl.load_workbook("formato.xlsx")
        sheet = book["Listas de chequeo SVCSP"]
        sheet.cell(7, 4).value = self.nombre
        sheet.cell(7, 8).value = self.localidad
        sheet.cell(9, 4).value = self.codigo

        book.save(f"{self.carpeta}/formato.xlsx")
        book.close()
