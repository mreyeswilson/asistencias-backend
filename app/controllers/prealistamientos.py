from openpyxl import load_workbook
from models.asistencia import Asistencia
from models.unidad import Unidad
from models.prealistamiento import Prealistamiento as PrealistamientoModel
import os
from db import db
from datetime import datetime, timedelta
from enums import MONTHS
import shutil

class Prealistamiento:

    folders = []
    month = MONTHS[datetime.now().strftime("%B")]
    filename = "LC_SIVIGILA_IAAS_CRONICOS_"

    def generar(self, filename, original_filename):

        prealistamiento = PrealistamientoModel(
            nombre=original_filename,
            fecha_prealistamiento=datetime.now(),
            archivo_zip=""
        )

        db.session.add(prealistamiento)

        if os.path.exists("tmp/prealistamientos"):
            shutil.rmtree("tmp/prealistamientos")

        wb = load_workbook(f"tmp/{filename}")
        
        ws = wb[wb.sheetnames[0]]
        asistencias = []
        for i in range(1, ws.max_row):
            codigo = str(ws.cell(i + 1, 4).value).strip()
            fecha_asistencia = str(ws.cell(i + 1, 6).value).split(" ")[0]
            hora_asistencia = ws.cell(i + 1, 7).value

            if hora_asistencia is None:
                hora_asistencia = "00:00:00"

            unidad = db.session.query(Unidad).filter_by(codigo=codigo).first()
            if unidad is None:
                asistencias.append(codigo)
                break
            
            fecha_asistencia = datetime.strptime(f"{fecha_asistencia} {hora_asistencia}", "%Y-%m-%d %H:%M:%S")
            fecha_prealistamiento = fecha_asistencia - timedelta(days=2)
            epidemiologo = ws.cell(i + 1, 8).value
            tecnico = unidad.localidad.tecnico

            asistencia = Asistencia(
                fecha_asistencia=fecha_asistencia,
                fecha_prealistamiento=fecha_prealistamiento,
                estado="pendiente",
                epidemiologo=epidemiologo,
                unidad_id=unidad.id,
                prealistamiento_id=prealistamiento.id
            )

            self.folders.append({
                "unidad": unidad,
                "asistencia": asistencia,
                "epidemiologo": epidemiologo,
                "tecnico": tecnico
            })

            db.session.add(asistencia)

        wb.close()
        
        if len(asistencias) > 0:
            db.session.rollback()
            return asistencias
        
        db.session.commit()
        zip_file = self.__generate_files(original_filename)
        os.unlink(f"tmp/{filename}")
        return zip_file

    def __generate_files(self, ori_filename):
        
        
        folder_path = f"storage/{self.month}/{ori_filename}"
        for f in self.folders:
            if os.path.exists(folder_path):
                os.makedirs(folder_path, exist_ok=True)

            pree_path = f"{folder_path}/{f['unidad'].nombre}"
            if not os.path.exists(pree_path):
                os.makedirs(pree_path, exist_ok=True)

            book = self.__fill_preenlistment_file(f)
            book.save(f"{pree_path}/{self.filename}{f['unidad'].nombre}.xlsx")
            book.close()
        # self.__zip_files(folder_path, ori_filename)
        # return f"{ori_filename}.zip"
    
    def __fill_preenlistment_file(self, f):
        wb = load_workbook("files/formato.xlsx")
        ws = wb["Listas de chequeo SVCSP"]

        ws["D9"] = f["unidad"].codigo
        ws["H7"] = f["unidad"].localidad.nombre
        ws["H9"] = datetime.now().strftime("%d/%m/%Y")

        ws = wb["Formato SIVIGILA"]
        ws["M17"] = f["epidemiologo"]
        ws["M15"] = f["tecnico"]
        ws["K11"] = ""
        ws["K13"] = ""
        ws["M13"] = ""

        return wb
    
    def __zip_files(self, path, ori_filename):
        print(path, ori_filename)
        dest = f"storage/{ori_filename}"
        shutil.make_archive(dest, 'zip', path)


prealistamiento = Prealistamiento()